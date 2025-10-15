import streamlit as st
import math
import os
import shutil
import openpyxl
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils import get_column_letter
import pandas as pd
import io
import base64
from datetime import datetime

# 设置页面配置
st.set_page_config(
    page_title="直流系统计算软件",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

class DCLoadCalculator:
    def __init__(self):
        # 初始化Kc值表 - 只使用1.85V放电终止电压
        self.kc_values_185 = {
            '5s': 1.34,
            '1min': 1.24,
            '29min': 0.8,
            '0.5h': 0.78,
            '59min': 0.558,
            '1.0h': 0.54,
            '89min': 0.432,
            '1.5h': 0.428,
            '119min': 0.347,
            '2.0h': 0.344,
            '179min': 0.263,
            '3.0h': 0.262,
            '4.0h': 0.214,
            '5.0h': 0.18,
            '6.0h': 0.157,
            '7.0h': 0.14,
            '479min': 0.123,
            '8.0h': 0.123
        }
        
        self.loads_data = []

    def calculate_current(self, capacity, load_factor):
        """计算电流：容量(kW) * 1000 * 负荷系数 / 220"""
        return capacity * 1000 * load_factor / 220

    def calculate_statistics(self):
        """计算电流统计 - 严格按照表格中的公式"""
        stats = {
            'I0': sum(load['frequent_current'] for load in self.loads_data),
            'I1': sum(load['cho_current'] for load in self.loads_data),
            'I2': sum(load['stage1_current'] for load in self.loads_data),
            'I3': sum(load['stage2_current'] for load in self.loads_data),
            'I4': sum(load['stage3_current'] for load in self.loads_data),
            'I5': sum(load['stage4_current'] for load in self.loads_data),
            'IR': sum(load['random_current'] for load in self.loads_data)
        }
        return stats

    def calculate_capacity(self, stats):
        """计算容量 - 严格按照表格中的公式和取值"""
        kc = self.kc_values_185
        
        capacity_calc = {}
        
        # 初期（1min）容量计算
        capacity_calc['initial'] = 1.4 * (stats['I1'] / kc['1min'])
        
        # 持续0.5h容量计算
        capacity_calc['stage1'] = 1.4 * ((stats['I1'] / kc['2.0h']) + 
                                        ((stats['I2'] - stats['I1']) / kc['29min']))
        
        # 持续1h容量计算
        capacity_calc['stage2'] = 1.4 * ((stats['I1'] / kc['2.0h']) + 
                                        ((stats['I2'] - stats['I1']) / kc['59min']) + 
                                        ((stats['I3'] - stats['I2']) / kc['0.5h']))
        
        # 持续2h容量计算
        capacity_calc['stage3'] = 1.4 * ((stats['I1'] / kc['2.0h']) + 
                                        ((stats['I2'] - stats['I1']) / kc['119min']) + 
                                        ((stats['I3'] - stats['I2']) / kc['1.5h']) + 
                                        ((stats['I4'] - stats['I3']) / kc['1.0h']))
         
        # 持续4h容量计算
        capacity_calc['stage4'] = 1.4 * ((stats['I1'] / kc['4.0h']) + 
                                        ((stats['I2'] - stats['I1']) / kc['4.0h']) + 
                                        ((stats['I5'] - stats['I4']) / kc['2.0h']))

        # 随机负荷容量计算
        capacity_calc['random'] = stats['IR'] / kc['5s']
        
        return capacity_calc

    def calculate_combined_load(self, capacity_calc):
        """计算叠加随机负荷 - 严格按照表格中的公式"""
        combined = {
            'initial': capacity_calc['initial'] + capacity_calc['random'],
            'stage1': capacity_calc['stage1'] + capacity_calc['random'],
            'stage2': capacity_calc['stage2'] + capacity_calc['random'],
            'stage3': capacity_calc['stage3'] + capacity_calc['random'],
            'stage4': capacity_calc['stage4'] + capacity_calc['random']
        }
        return combined

    def calculate_final_capacity(self, combined_load):
        """计算最终容量取值（向上取整）"""
        max_capacity = max(combined_load.values())
        return math.ceil(max_capacity)

class BatteryCountCalculator:
    """蓄电池个数计算器"""
    def __init__(self):
        self.default_un = 220  # 直流电源系统标称电压（V）
        self.default_uf = 2.23  # 单体蓄电池浮充电电压（V）
    
    def calculate_battery_count(self, un, uf):
        """计算蓄电池个数：n = (Un / Uf) * 1.05，然后向上取整"""
        n = (un / uf) * 1.05
        return math.ceil(n)
    
    def calculate_with_inputs(self, un_input, uf_input):
        """根据输入计算蓄电池个数，处理输入验证"""
        try:
            un = float(un_input)
            uf = float(uf_input)
            
            if un <= 0 or uf <= 0:
                return None, "电压值必须大于0"
            
            battery_count = self.calculate_battery_count(un, uf)
            calculation_process = f"计算过程:\n"
            calculation_process += f"n = (Un / Uf) × 1.05\n"
            calculation_process += f"  = ({un} / {uf}) × 1.05\n"
            calculation_process += f"  = {un/uf:.4f} × 1.05\n"
            calculation_process += f"  = {(un/uf)*1.05:.4f}\n"
            calculation_process += f"向上取整 = {battery_count}"
            
            return battery_count, calculation_process
            
        except ValueError:
            return None, "请输入有效的数字"
        except ZeroDivisionError:
            return None, "Uf（浮充电电压）不能为0"

class HighFrequencyPowerModuleCalculator:
    """高频开关电源模块选择数量计算器"""
    def __init__(self):
        self.default_frequent_current = 27.27  # 默认经常负荷电流 (A)
        self.default_module_current = 20  # 默认单个模块额定电流 (A)
    
    def calculate_module_count(self, battery_capacity, frequent_current, module_current):
        """
        计算高频开关电源模块选择数量
        步骤：
        1. 计算电流 = 1.25 × (蓄电池容量 ÷ 10) + 经常负荷电流
        2. n1 = 计算电流 ÷ 单个模块额定电流，向上取整
        3. n2 = 1 (当n1 <= 6) 或 2 (当n1 >= 7)
        4. n = n1 + n2
        """
        try:
            # 1. 计算电流
            calc_current = 1.25 * (battery_capacity / 10) + frequent_current
            
            # 2. 计算n1（基本模块数量）
            n1 = math.ceil(calc_current / module_current)
            
            # 3. 计算n2（附加模块数量）
            n2 = 1 if n1 <= 6 else 2
            
            # 4. 计算总模块数量
            total_modules = n1 + n2
            
            # 生成计算过程说明
            calculation_process = f"计算过程:\n"
            calculation_process += f"1. 计算电流 = 1.25 × (蓄电池容量 ÷ 10) + 经常负荷电流\n"
            calculation_process += f"   = 1.25 × ({battery_capacity} ÷ 10) + {frequent_current}\n"
            calculation_process += f"   = 1.25 × {battery_capacity/10:.2f} + {frequent_current}\n"
            calculation_process += f"   = {1.25*(battery_capacity/10):.2f} + {frequent_current}\n"
            calculation_process += f"   = {calc_current:.2f} A\n\n"
            calculation_process += f"2. n1 = 计算电流 ÷ 单个模块额定电流 (向上取整)\n"
            calculation_process += f"   = {calc_current:.2f} ÷ {module_current}\n"
            calculation_process += f"   = {calc_current/module_current:.2f}\n"
            calculation_process += f"   向上取整 = {n1}\n\n"
            calculation_process += f"3. n2 = 附加模块数量\n"
            calculation_process += f"   n1 = {n1}, 因此n2 = {n2}\n\n"
            calculation_process += f"4. 总模块数量 n = n1 + n2\n"
            calculation_process += f"   = {n1} + {n2}\n"
            calculation_process += f"   = {total_modules}"
            
            return {
                'calc_current': calc_current,
                'n1': n1,
                'n2': n2,
                'total_modules': total_modules,
                'process': calculation_process
            }
            
        except Exception as e:
            return None

def create_download_link(file_data, file_name, link_text):
    """创建文件下载链接"""
    b64 = base64.b64encode(file_data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{file_name}">{link_text}</a>'
    return href

def get_file_downloads():
    """获取可下载的文件列表和内容"""
    downloads = []
    
    # 文件大小限制（以字节为单位）
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    
    # 文件1: 直流负荷统计.docx
    try:
        if os.path.exists("直流负荷统计.docx"):
            file_size = os.path.getsize("直流负荷统计.docx")
            if file_size > MAX_FILE_SIZE:
                st.warning(f"直流负荷统计.docx 文件较大 ({file_size/1024/1024:.1f}MB)，下载可能需要较长时间")
            
            with open("直流负荷统计.docx", "rb") as f:
                docx_content = f.read()
            docx_description = f"直流负荷统计文档，包含详细的负荷统计说明和表格 ({file_size/1024:.1f}KB)"
        else:
            # 如果文件不存在，创建一个小型示例文件
            docx_content = b"DC Load Statistics Document - Sample Content"
            docx_description = "直流负荷统计文档，包含详细的负荷统计说明和表格 (示例文件)"
    except Exception as e:
        st.error(f"加载直流负荷统计.docx文件时出错: {str(e)}")
        docx_content = b"Error loading file"
        docx_description = "文件加载出错"
    
    downloads.append({
        "name": "直流负荷统计.docx",
        "content": docx_content,
        "description": docx_description
    })
    
    # 文件2: 直流负荷统计.xlsx
    try:
        if os.path.exists("直流负荷统计.xlsx"):
            file_size = os.path.getsize("直流负荷统计.xlsx")
            if file_size > MAX_FILE_SIZE:
                st.warning(f"直流负荷统计.xlsx 文件较大 ({file_size/1024/1024:.1f}MB)，下载可能需要较长时间")
            
            with open("直流负荷统计.xlsx", "rb") as f:
                excel_content = f.read()
            excel_description = f"直流负荷统计Excel表格，包含负荷数据和计算公式 ({file_size/1024:.1f}KB)"
        else:
            # 如果文件不存在，创建示例Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "直流负荷统计"
            
            # 添加标题
            ws['A1'] = "直流负荷统计表"
            ws['A1'].font = Font(size=14, bold=True)
            
            # 添加表头
            headers = ["序号", "负荷名称", "容量(kW)", "负荷系数", "计算电流(A)"]
            for i, header in enumerate(headers):
                ws.cell(row=3, column=i+1, value=header)
                ws.cell(row=3, column=i+1).font = Font(bold=True)
            
            # 添加示例数据
            example_data = [
                ["控制、保护、继电器", 10, 0.6, 27.27],
                ["断路器跳闸", 3.6, 0.6, 9.82],
                ["UPS电源", 15, 0.6, 40.91],
            ]
            
            for i, data in enumerate(example_data):
                ws.cell(row=4+i, column=1, value=i+1)
                for j, value in enumerate(data):
                    ws.cell(row=4+i, column=j+2, value=value)
            
            # 保存到字节流
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_content = excel_buffer.getvalue()
            excel_description = "直流负荷统计Excel表格，包含负荷数据和计算公式 (示例文件)"
    except Exception as e:
        st.error(f"加载直流负荷统计.xlsx文件时出错: {str(e)}")
        excel_content = b"Error loading file"
        excel_description = "文件加载出错"
    
    downloads.append({
        "name": "直流负荷统计.xlsx",
        "content": excel_content,
        "description": excel_description
    })
    
    # 文件3: 直流负荷统计.exe - 使用外部下载链接
    exe_external_url = "https://wwya.lanzoue.com/i3O7s38g9pyb"
    exe_password = "1tn7"
    
    downloads.append({
        "name": "直流负荷统计.exe",
        "external_url": exe_external_url,
        "password": exe_password,
        "description": "直流负荷统计桌面应用程序，可在Windows系统上独立运行 (通过外部链接下载)，建议关闭杀毒软件，以防止其被杀毒软件清除"
    })
    
    return downloads

def main():
    # 初始化计算器
    dc_calculator = DCLoadCalculator()
    battery_calculator = BatteryCountCalculator()
    hf_power_calculator = HighFrequencyPowerModuleCalculator()
    
    # 页面标题
    st.title("⚡ 直流系统计算软件")
    st.caption("参考：《DLT 5044-2014 电力工程直流电源系统设计技术规程》")
    st.markdown("---")
    
    # 创建标签页
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 直流负荷计算", 
        "🔋 蓄电池个数计算", 
        "🔌 高频开关电源模块选择",
        "📥 文件下载"
    ])
    
    # 标签页1: 直流负荷计算
    with tab1:
        st.header("直流负荷计算")
        
        # 初始化session state
        if 'loads_data' not in st.session_state:
            st.session_state.loads_data = []
        
        # 输入表单
        with st.form("load_input_form"):
            col1, col2, col3 = st.columns(3)
            
            with col1:
                name = st.text_input("负荷名称", placeholder="输入负荷名称")
            
            with col2:
                capacity = st.number_input("容量(kW)", min_value=0.0, max_value=1000.0, value=10.0, step=0.1)
            
            with col3:
                load_factor = st.number_input("负荷系数", min_value=0.0, max_value=1.0, value=0.6, step=0.1)
            
            # 负荷阶段选择
            st.subheader("负荷阶段选择")
            stage_cols = st.columns(7)
            
            with stage_cols[0]:
                frequent = st.checkbox("经常负荷", value=True)
            with stage_cols[1]:
                cho = st.checkbox("初期(1min)", value=True)
            with stage_cols[2]:
                stage1 = st.checkbox("0.5h", value=True)
            with stage_cols[3]:
                stage2 = st.checkbox("1h", value=True)
            with stage_cols[4]:
                stage3 = st.checkbox("2h", value=True)
            with stage_cols[5]:
                stage4 = st.checkbox("4h", value=False)
            with stage_cols[6]:
                random = st.checkbox("随机(5s)", value=False)
            
            # 提交按钮
            submitted = st.form_submit_button("添加负荷")
            
            if submitted:
                if not name:
                    st.error("请输入负荷名称")
                else:
                    # 计算电流
                    current = dc_calculator.calculate_current(capacity, load_factor)
                    
                    # 添加到数据
                    load_data = {
                        'name': name,
                        'capacity': capacity,
                        'load_factor': load_factor,
                        'calc_current': current,
                        'frequent_current': current if frequent else 0,
                        'cho_current': current if cho else 0,
                        'stage1_current': current if stage1 else 0,
                        'stage2_current': current if stage2 else 0,
                        'stage3_current': current if stage3 else 0,
                        'stage4_current': current if stage4 else 0,
                        'random_current': current if random else 0
                    }
                    st.session_state.loads_data.append(load_data)
                    st.success(f"负荷 '{name}' 添加成功!")
        
        # 示例数据按钮
        if st.button("加载示例数据"):
            example_loads = [
                ("控制、保护、继电器", 10, 0.6, True, True, True, True, True, False, False),
                ("断路器跳闸", 3.6, 0.6, False, True, False, False, False, False, False),
                ("断路器自投", 1.8, 1, False, False, False, False, False, False, True),
                ("断路器合闸", 1.8, 1, False, False, False, False, False, False, True),
                ("UPS电源", 15, 0.6, False, True, True, True, True, False, False),
                ("全场事故照明负荷", 3, 1, False, True, True, True, True, False, False),
                ("DC/DC变换装置", 3, 0.8, False, False, False, False, False, True, False),
            ]
            
            st.session_state.loads_data = []
            for load in example_loads:
                name, capacity, load_factor, frequent, cho, stage1, stage2, stage3, stage4, random = load
                current = dc_calculator.calculate_current(capacity, load_factor)
                load_data = {
                    'name': name,
                    'capacity': capacity,
                    'load_factor': load_factor,
                    'calc_current': current,
                    'frequent_current': current if frequent else 0,
                    'cho_current': current if cho else 0,
                    'stage1_current': current if stage1 else 0,
                    'stage2_current': current if stage2 else 0,
                    'stage3_current': current if stage3 else 0,
                    'stage4_current': current if stage4 else 0,
                    'random_current': current if random else 0
                }
                st.session_state.loads_data.append(load_data)
            st.success("示例数据加载成功!")
        
        # 清空按钮
        if st.button("清空所有负荷"):
            st.session_state.loads_data = []
            st.success("所有负荷已清空!")
        
        # 显示负荷表格
        if st.session_state.loads_data:
            st.subheader("负荷列表")
            
            # 准备表格数据
            table_data = []
            for i, load in enumerate(st.session_state.loads_data):
                row = [
                    i + 1,
                    load['name'],
                    f"{load['capacity']:.2f}",
                    f"{load['load_factor']:.2f}",
                    f"{load['calc_current']:.2f}",
                    "是" if load['frequent_current'] > 0 else "否",
                    "是" if load['cho_current'] > 0 else "否",
                    "是" if load['stage1_current'] > 0 else "否",
                    "是" if load['stage2_current'] > 0 else "否",
                    "是" if load['stage3_current'] > 0 else "否",
                    "是" if load['stage4_current'] > 0 else "否",
                    "是" if load['random_current'] > 0 else "否"
                ]
                table_data.append(row)
            
            # 显示表格
            df = pd.DataFrame(
                table_data,
                columns=['序号', '负荷名称', '容量(kW)', '负荷系数', '计算电流(A)', 
                        '经常负荷', '初期', '0.5h', '1h', '2h', '4h', '随机']
            )
            st.dataframe(df, use_container_width=True)
            
            # 计算按钮
            if st.button("开始计算"):
                dc_calculator.loads_data = st.session_state.loads_data
                
                try:
                    # 计算统计
                    stats = dc_calculator.calculate_statistics()
                    
                    # 计算容量
                    capacity_calc = dc_calculator.calculate_capacity(stats)
                    
                    # 计算叠加负荷
                    combined_load = dc_calculator.calculate_combined_load(capacity_calc)
                    
                    # 计算最终容量
                    final_capacity = dc_calculator.calculate_final_capacity(combined_load)
                    
                    # 显示结果
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.subheader("电流统计 (A)")
                        st.text(f"I0 (经常负荷):    {stats['I0']:.2f} A")
                        st.text(f"I1 (初期 1min):   {stats['I1']:.2f} A")
                        st.text(f"I2 (0.5h):        {stats['I2']:.2f} A")
                        st.text(f"I3 (1h):          {stats['I3']:.2f} A")
                        st.text(f"I4 (2h):          {stats['I4']:.2f} A")
                        st.text(f"I5 (4h):          {stats['I5']:.2f} A")
                        st.text(f"IR (随机 5s):     {stats['IR']:.2f} A")
                    
                    with col2:
                        st.subheader("容量计算 (Ah)")
                        st.text(f"初期 (1min):      {capacity_calc['initial']:.2f} Ah")
                        st.text(f"持续0.5h:         {capacity_calc['stage1']:.2f} Ah")
                        st.text(f"持续1h:           {capacity_calc['stage2']:.2f} Ah")
                        st.text(f"持续2h:           {capacity_calc['stage3']:.2f} Ah")
                        st.text(f"持续4h:           {capacity_calc['stage4']:.2f} Ah")
                        st.text(f"随机负荷:         {capacity_calc['random']:.2f} Ah")
                    
                    with col3:
                        st.subheader("叠加随机负荷 (Ah)")
                        st.text(f"初期+随机:        {combined_load['initial']:.2f} Ah")
                        st.text(f"0.5h+随机:        {combined_load['stage1']:.2f} Ah")
                        st.text(f"1h+随机:          {combined_load['stage2']:.2f} Ah")
                        st.text(f"2h+随机:          {combined_load['stage3']:.2f} Ah")
                        st.text(f"4h+随机:          {combined_load['stage4']:.2f} Ah")
                    
                    st.success(f"最终计算容量: 各设计取值不一，以上结果可供参考，最终取值以个人取值为准。")
                    
                except Exception as e:
                    st.error(f"计算过程中发生错误: {str(e)}")
        
        else:
            st.info("暂无负荷数据，请添加负荷或加载示例数据")
    
    # 标签页2: 蓄电池个数计算
    with tab2:
        st.header("蓄电池个数计算")
        
        st.markdown("""
        **计算公式**: n = (Un / Uf) × 1.05，然后向上取整  
        其中：  
        - n   —— 蓄电池个数  
        - Un  —— 直流电源系统标称电压（V）  
        - Uf  —— 单体蓄电池浮充电电压（V）
        """)
        
        col1, col2 = st.columns(2)
        
        with col1:
            un = st.number_input("直流电源系统标称电压 Un (V)", 
                                min_value=0.0, value=220.0, step=1.0)
        
        with col2:
            uf = st.number_input("单体蓄电池浮充电电压 Uf (V)", 
                                min_value=0.0, value=2.23, step=0.01)
        
        if st.button("计算蓄电池个数"):
            battery_count, process_text = battery_calculator.calculate_with_inputs(str(un), str(uf))
            
            if battery_count is not None:
                st.success(f"蓄电池个数: {battery_count} 个")
                st.text_area("计算过程", process_text, height=200)
            else:
                st.error(process_text)
    
    # 标签页3: 高频开关电源模块选择
    with tab3:
        st.header("高频开关电源模块选择数量计算")
        
        st.markdown("""
        **计算步骤**:  
        1. 计算电流 = 1.25 × (蓄电池容量 ÷ 10) + 经常负荷电流  
        2. n1 = 计算电流 ÷ 单个模块额定电流 (向上取整)  
        3. n2 = 1 (当n1 ≤ 6) 或 2 (当n1 ≥ 7)  
        4. 总模块数量 n = n1 + n2  
        
        其中：  
        - n1   —— 基本模块数量  
        - n2   —— 附加模块数量  
        - Imo  —— 单个模块额定电流  
        - Ijc  —— 经常负荷电流
        """)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            battery_capacity = st.number_input("蓄电池容量 (Ah)", 
                                             min_value=0.0, value=400.0, step=1.0)
        
        with col2:
            frequent_current = st.number_input("经常负荷电流 Ijc (A)", 
                                             min_value=0.0, value=27.27, step=0.1)
        
        with col3:
            module_current = st.number_input("单个模块额定电流 Imo (A)", 
                                           min_value=0.0, value=20.0, step=1.0)
        
        if st.button("计算模块数量"):
            result = hf_power_calculator.calculate_module_count(
                battery_capacity, frequent_current, module_current
            )
            
            if result:
                st.success(
                    f"高频开关电源模块选择数量: {result['total_modules']} 个  "
                    f"(n1 = {result['n1']}, n2 = {result['n2']})"
                )
                st.text_area("计算过程", result['process'], height=600)
            else:
                st.error("计算过程中发生错误")

    # 标签页4: 文件下载
    with tab4:
        st.header("📥 文件下载中心")
        st.markdown("---")
        
        st.info("""
        在这里您可以下载直流负荷统计相关的文件和工具。所有文件都经过安全检查，可以直接下载使用。
        """)
        
        # 获取下载文件列表
        downloads = get_file_downloads()
        
        # 显示文件下载卡片
        for i, file_info in enumerate(downloads):
            with st.container():
                col1, col2 = st.columns([3, 1])
                
                with col1:
                    st.subheader(f"📄 {file_info['name']}")
                    st.write(file_info['description'])
                    
                    # 显示文件信息
                    if "content" in file_info:
                        file_size = len(file_info['content'])
                        st.caption(f"文件大小: {file_size / 1024:.1f} KB | 更新日期: {datetime.now().strftime('%Y-%m-%d')}")
                    elif "external_url" in file_info:
                        st.caption(f"外部下载链接 | 更新日期: {datetime.now().strftime('%Y-%m-%d')}")
                
                with col2:
                    # 创建下载链接
                    if "external_url" in file_info:
                        # 外部链接文件
                        external_url = file_info["external_url"]
                        password = file_info.get("password", "")
                        
                        # 显示密码信息
                        if password:
                            st.info(f"提取密码: {password}")
                        
                        # 创建外部链接按钮
                        st.markdown(f'<a href="{external_url}" target="_blank" style="text-decoration: none;"><button style="background-color: #4CAF50; color: white; padding: 10px 20px; border: none; border-radius: 4px; cursor: pointer;">🌐 外部下载</button></a>', 
                                   unsafe_allow_html=True)
                    else:
                        # 本地文件
                        download_link = create_download_link(
                            file_info['content'], 
                            file_info['name'], 
                            "📥 下载文件"
                        )
                        st.markdown(download_link, unsafe_allow_html=True)
                
                # 添加分隔线（除了最后一个文件）
                if i < len(downloads) - 1:
                    st.markdown("---")
        
        # 添加使用说明
        st.markdown("---")
        st.subheader("使用说明")
        
        with st.expander("文件使用指南", expanded=False):
            st.markdown("""
            ### 📋 文件说明
            
            **1. 直流负荷统计.docx**
            - 完善的表格文件
            - 适合打印和文档归档
            
            **2. 直流负荷统计.xlsx**
            - 包含负荷数据表格和计算公式（可使用“直流负荷统计.exe”导出负荷计算表格）
            - 可以直接编辑和使用
            
            **3. 直流负荷统计.exe**
            - 独立的桌面应用程序
            - 无需安装，直接运行
            - 包含所有计算功能
            - 通过外部链接下载，提取密码: 22rw
            
            ### 🔒 安全提示
            - 所有文件都经过安全检查
            - 下载后建议进行病毒扫描
            - 如有问题请联系系统管理员
            
            ### 📞 技术支持
            如有任何问题或需要帮助，请联系技术支持团队。
            """)
        
        # 添加反馈部分
        with st.expander("问题反馈", expanded=False):
            feedback = st.text_area("如果您在使用过程中遇到问题或有改进建议，请告诉我们：")
            if st.button("提交反馈"):
                if feedback:
                    st.success("感谢您的反馈！我们会尽快处理。")
                else:
                    st.warning("请输入您的反馈内容。")

if __name__ == "__main__":
    main()